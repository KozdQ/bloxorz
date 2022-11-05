import os
import sys
import time
from multiprocessing import Process, Manager

import psutil
import pyautogui as pyautogui
from openpyxl import Workbook
from selenium import webdriver
from selenium.common import TimeoutException
from selenium.webdriver.chrome.service import Service
from chromedriver_py import binary_path
from selenium.webdriver.common.by import By

from block import Block
from engine import Engine
from func import Func
from map import Map
from stat_info import Stat

sys.setrecursionlimit(0x100000)

START_MAP = 1
END_MAP = 33
TIMELIMIT = 3  # seconds
STAT_CMD_MODE = False


def readInput(input_file):
    # open file
    with open(input_file) as f:
        # read first line
        map_row, map_col, row_start, col_start, numFunc = [int(x) for x in next(f).split()]

        # read list function's sequence
        func_seq = []
        for _ in range(numFunc):
            func_seq.append([int(x) for x in next(f).split()])

        # read map
        map_seq = []
        for _ in range(map_row):
            map_seq.append([int(x) for x in next(f).split()])

    return map_row, map_col, row_start, col_start, map_seq, func_seq


def mainAll(is_visual, engine_list):
    driver = None
    if is_visual:
        service_object = Service(binary_path)
        driver = webdriver.Chrome(service=service_object)
        driver.get("https://www.mathplayground.com/logic_bloxorz.html")
        driver.maximize_window()
    try:
        if is_visual:
            time.sleep(3)
            shadow_host = driver.find_element(By.TAG_NAME, 'ruffle-player')
            driver.execute_script(
                "arguments[0].scrollIntoView({'block':'center','inline':'center'})",
                shadow_host)
            shadow_root = shadow_host.shadow_root
            myElem = shadow_root.find_element(By.ID, 'play_button')
            if myElem.get_attribute("style") == "display: block;":
                myElem.click()
                time.sleep(12)
                nextStepProcess = True
            else:
                nextStepProcess = True
        else:
            nextStepProcess = True

        if nextStepProcess:

            startMain = START_MAP
            if is_visual:
                if START_MAP == 1:
                    pyautogui.click(607, 562)
                    time.sleep(2)
                    pyautogui.click(912, 688)
                    time.sleep(6)
                else:
                    passCode = [idx for idx in open("passcode/passcode.txt").readlines()[START_MAP - 1][18:-1]]

                    pyautogui.click(607, 582)
                    time.sleep(1)
                    pyautogui.press(passCode, interval=0.1)
                    pyautogui.click(650, 621)
                    time.sleep(6)

            endMain = END_MAP

            for idx in range(startMain, endMain + 1):

                if is_visual:
                    engine_list[idx].printer.doRoad(idxMap=idx)

                    engine_list[idx].printer.pressListKey()

                    if startMain == endMain:
                        time.sleep(3)
                    else:
                        time.sleep(6)

                else:
                    if STAT_CMD_MODE:
                        engine_list[idx].printer.stat.printStat(idx)
                    else:
                        engine_list[idx].printer.printRoad(idxMap=idx)
                        engine_list[idx].printer.printPressKey()

    except TimeoutException as ex:
        print(ex)
    finally:
        if is_visual:
            driver.quit()


def process_memory():
    process = psutil.Process(os.getpid())
    mem_info = process.memory_info()
    return mem_info.rss


def oneEngine(idx, alg):
    inputFile = "input/map-{num:02d}.txt".format(num=idx)

    mapRow, mapCol, rowStart, colStart, mapSeq, funcSeq = readInput(inputFile)

    # SET STAT
    DFSStat, BFSStat, ASTARStat = None, None, None
    if alg == "DFS":
        DFSStat = Stat("DFS Stat")
        DFSStat.startMemory = process_memory()
    elif alg == "BFS":
        BFSStat = Stat("BFS Stat")
        BFSStat.startMemory = process_memory()
    elif alg == "ASTAR":
        ASTARStat = Stat("ASTAR Stat")
        ASTARStat.startMemory = process_memory()

    # SET MAP
    srcMap = Map(mapSeq, mapRow, mapCol)

    # SET FUNCTION MAP
    func = Func(funcSeq)

    # SET BLOCK
    block = Block(rowStart, colStart, rowStart, colStart, "STANDING", None, srcMap)

    # SET ENGINE
    engine = Engine(srcMap, func, block)

    # RUN
    if alg == "DFS":
        engine.printer.stat = DFSStat
        engine.DFS()
    elif alg == "BFS":
        engine.printer.stat = BFSStat
        engine.BFS()
    elif alg == "ASTAR":
        engine.printer.stat = ASTARStat
        engine.PriorityDistanceFS()

    engine.printer.stat.processMemory = process_memory() - engine.printer.stat.startMemory

    return engine


def do(idx, alg, ret_dict):
    if alg == "DFS":
        algIdx = 1
    elif alg == "BFS":
        algIdx = 2
    elif alg == "ASTAR":
        algIdx = 3
    else:
        algIdx = 0
    ret_dict["{i}-{j:02d}".format(i=algIdx, j=idx)] = oneEngine(idx, alg)


def getList(alg, ret_dict):
    if alg == "DFS":
        algIdx = 1
    elif alg == "BFS":
        algIdx = 2
    elif alg == "ASTAR":
        algIdx = 3
    else:
        algIdx = 0
    thisList = []
    for jdx in range(34):
        thisList.append(ret_dict["{i}-{j:02d}".format(i=algIdx, j=jdx)])
    return thisList


def validate(alg, list_engine, s, e):
    if alg == "DFS":
        algIdx = 1
    elif alg == "BFS":
        algIdx = 2
    elif alg == "ASTAR":
        algIdx = 3
    else:
        algIdx = 0
    tmpBool = True
    for idx in range(s, e + 1):
        if list_engine[idx] is None:
            print("TIMEOUT at algIdx", algIdx, ", idx", idx)
            tmpBool = False
    return tmpBool


def benchmarkFunc(alg, list_engine):
    wb = Workbook()
    dest_filename = str(alg).lower() + '_benchmark.xlsx'
    ws1 = wb.active
    ws1.title = str(alg) + " benchmark"

    # NAME
    _ = ws1.cell(column=1, row=1, value="MAP")
    _ = ws1.cell(column=1, row=2, value="STATUS")
    _ = ws1.cell(column=1, row=3, value="START TIME")
    _ = ws1.cell(column=1, row=4, value="PROCESS TIME")
    _ = ws1.cell(column=1, row=5, value="REAL STEP")
    _ = ws1.cell(column=1, row=6, value="VIRTUAL STEP")
    _ = ws1.cell(column=1, row=7, value="PROCESS MEMORY")

    for col in range(1, 34):
        if list_engine[col]:
            tmp = list_engine[col].printer.stat
            status = tmp.status
            startTime = tmp.startTime
            runningTime = tmp.runningTime
            countStep = tmp.countStep
            virtualStep = tmp.virtualStep
            processMemory = tmp.processMemory
        else:
            status = "FAIL"
            startTime = "N/A"
            runningTime = "N/A"
            countStep = "N/A"
            virtualStep = "N/A"
            processMemory = "N/A"

        _ = ws1.cell(column=col + 1, row=1, value="{num:02d}".format(num=col))
        _ = ws1.cell(column=col + 1, row=2, value="{s}".format(s=str(status)))
        _ = ws1.cell(column=col + 1, row=3, value="{s}".format(s=str(startTime)))
        _ = ws1.cell(column=col + 1, row=4, value="{s}".format(s=str(runningTime)))
        _ = ws1.cell(column=col + 1, row=5, value="{s}".format(s=str(countStep)))
        _ = ws1.cell(column=col + 1, row=6, value="{s}".format(s=str(virtualStep)))
        _ = ws1.cell(column=col + 1, row=7, value="{s}".format(s=str(processMemory)))

    wb.save(dest_filename)


if __name__ == '__main__':
    mapMode = sys.argv[1]
    isVisual = (int(sys.argv[2]) == 1)
    algorithm = sys.argv[3]
    benchmark = (int(sys.argv[4]) == 1)
    STAT_CMD_MODE = (int(sys.argv[5]) == 1)

    manager = Manager()
    return_dict = manager.dict()
    for i in range(4):
        for j in range(34):
            return_dict["{i}-{j:02d}".format(i=i, j=j)] = None

    mapMode = mapMode.split(sep="-")
    if len(mapMode) == 1:
        if mapMode[0] != "all":
            START_MAP = int(mapMode[0])
            END_MAP = START_MAP

        else:
            START_MAP = 1

    elif len(mapMode) == 2:
        START_MAP = int(mapMode[0])
        END_MAP = int(mapMode[1])

    else:
        print("ERROR MAP MODE PARAM")
        sys.exit()

    start = START_MAP
    end = END_MAP

    for i in range(start, end + 1):
        if algorithm in ["DFS", "BFS", "ASTAR"]:
            thread = Process(target=do, args=(i, algorithm, return_dict))
            thread.start()
            thread.join(timeout=TIMELIMIT)
            thread.terminate()
        elif algorithm == "BEST" and not isVisual:
            pass

    listEngine = getList(algorithm, return_dict)

    ADD_TIME = 20
    if benchmark:
        ADD_TIME = 60

    if not validate(algorithm, listEngine, start, end):
        if benchmark:
            print("NONE and retry")
        else:
            print("NONE, retry and maybe replace with DFS")
        for i in range(start, end + 1):
            if listEngine[i] is None:
                if algorithm == "DFS":
                    algorithmIdx = 1
                elif algorithm == "BFS":
                    algorithmIdx = 2
                elif algorithm == "ASTAR":
                    algorithmIdx = 3
                else:
                    algorithmIdx = 0

                # retry with 3 + ADD_TIME secs
                thread = Process(target=do, args=(i, algorithm, return_dict))
                thread.start()
                thread.join(timeout=TIMELIMIT + ADD_TIME)
                thread.terminate()

                if return_dict["{i}-{j:02d}".format(i=algorithmIdx, j=i)] is None:
                    print("Retry {num:02d} -> None".format(num=i))
                    if not benchmark:
                        print("Retry {num:02d} -> None -> Change algorithm to DFS".format(num=i))
                        # replace algorithm with DFS in 3 + 60 secs
                        thread = Process(target=do, args=(i, "DFS", return_dict))
                        thread.start()
                        thread.join(timeout=TIMELIMIT + 60)
                        thread.terminate()

                        if return_dict["{i}-{j:02d}".format(i=1, j=i)] is None:
                            print("Retry {num:02d} -> None -> Change algorithm to DFS -> None".format(num=i))
                        else:
                            print("Retry {num:02d} -> None -> Change algorithm to DFS -> Success".format(num=i))
                            listEngine[i] = return_dict["{i}-{j:02d}".format(i=1, j=i)]
                else:
                    print("Retry {num:02d} -> Success".format(num=i))
                    listEngine[i] = return_dict["{i}-{j:02d}".format(i=algorithmIdx, j=i)]
    if benchmark:
        benchmarkFunc(algorithm, listEngine)
        print("BENCHMARK " + str(algorithm))
    else:
        mainAll(isVisual, listEngine)
    sys.exit()
